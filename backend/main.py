"""
ProjektLLM — FastAPI backend
Serves the API + static frontend build.
"""

from __future__ import annotations

import asyncio
import json
import os
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import AsyncGenerator

from fastapi import FastAPI, HTTPException, UploadFile, File as FastAPIFile, Depends
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from sqlalchemy import select, delete, func, or_, and_, update as sql_update
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database import (
    init_db, get_db,
    Project, Chat, Message, File, AppSetting, User, ProjectShare,
    get_setting, set_setting,
)
from backend.providers import LLMConfig, ChatMessage, CompletionRequest, ToolDef, get_provider
from backend.routers import auth_router, documents_router, skills_router
from backend.routers.documents import _extract_docx_text, _extract_pdf_text
from backend.routers.documents import extract_text_from_file
from backend.auth import require_user, require_admin

# Disable auth for development — set to "1" to skip JWT checks
NO_AUTH = os.environ.get("PROJEKTLLM_NO_AUTH", "0") == "1"

# Auth dependencies — resolved once at module load to avoid FastAPI mistaking
# a plain `None` default for a body parameter when NO_AUTH is enabled.
if NO_AUTH:
    async def _noauth_user():
        return None
    _current_user = Depends(_noauth_user)
    _current_admin = Depends(_noauth_user)
else:
    _current_user = Depends(require_user)
    _current_admin = Depends(require_admin)


# ---------------------------------------------------------------------------
# App lifecycle
# ---------------------------------------------------------------------------

STATIC_DIR = Path(__file__).parent / "static"


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    yield


app = FastAPI(title="ProjektLLM", lifespan=lifespan)
app.include_router(auth_router)
app.include_router(documents_router)
app.include_router(skills_router)

# Serve built frontend
if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(STATIC_DIR / "assets")), name="assets")
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class ProjectOut(BaseModel):
    id: str
    name: str
    color: str
    glyph: str
    description: str
    instructions: str
    created_at: str
    updated_at: str
    owner_id: str | None = None
    chats: list = []
    files: list = []


class ProjectPatch(BaseModel):
    name: str | None = None
    color: str | None = None
    glyph: str | None = None
    description: str | None = None
    instructions: str | None = None


class ChatOut(BaseModel):
    id: str
    project_id: str
    title: str
    pinned: bool
    created_at: str
    updated_at: str
    message_count: int = 0


class ChatPatch(BaseModel):
    title: str | None = None
    pinned: bool | None = None


class MessageOut(BaseModel):
    id: str
    chat_id: str
    role: str
    content: str
    model: str | None = None
    citations: dict | None = None
    reasoning: str | None = None
    created_at: str


class FileOut(BaseModel):
    id: str
    project_id: str
    name: str
    kind: str
    size: str
    tokens: int
    added_at: str


class SendMessageBody(BaseModel):
    content: str
    model: str | None = None
    artifact_mode: bool = False
    web_search_context: str | None = None
    max_tokens: int | None = None
    temperature: float | None = None


class ArtifactBody(BaseModel):
    name: str
    content: str


class WebSearchBody(BaseModel):
    query: str


class SettingsBody(BaseModel):
    provider: str | None = None
    ollama_base: str | None = None
    ollama_model: str | None = None
    openai_base: str | None = None
    openai_key: str | None = None
    openai_model: str | None = None
    deepseek_key: str | None = None
    deepseek_model: str | None = None
    # Profile
    display_name: str | None = None
    avatar_initials: str | None = None
    llm_call_name: str | None = None
    # Generation
    temperature: float | None = None
    max_tokens: int | None = None
    stream: bool | None = None
    # Appearance
    chat_font_size: int | None = None
    enter_to_send: bool | None = None
    # Memory
    memories: list | None = None


# Sharing schemas
class ShareBody(BaseModel):
    username: str
    permission: str = "view"  # "view" | "edit"


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

async def _get_current_user_id(current_user: dict) -> str | None:
    return current_user.get("user_id") if current_user else None


async def _can_access_project(pid: str, user_id: str, db: AsyncSession) -> bool:
    """Check if user owns or has a share on the project."""
    project = await db.get(Project, pid)
    if not project:
        return False
    if project.owner_id == user_id:
        return True
    share = await db.execute(
        select(ProjectShare).where(
            ProjectShare.project_id == pid,
            ProjectShare.user_id == user_id,
        )
    )
    return share.scalar_one_or_none() is not None


async def _can_edit_project(pid: str, user_id: str, db: AsyncSession) -> bool:
    """Check if user can edit a project (owner or edit share)."""
    project = await db.get(Project, pid)
    if not project:
        return False
    if project.owner_id == user_id:
        return True
    share = await db.execute(
        select(ProjectShare).where(
            ProjectShare.project_id == pid,
            ProjectShare.user_id == user_id,
            ProjectShare.permission == "edit",
        )
    )
    return share.scalar_one_or_none() is not None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _id() -> str:
    return uuid.uuid4().hex[:12]


async def _chat_dict(chat: Chat, db: AsyncSession) -> dict:
    msg_count = await db.scalar(
        select(func.count(Message.id)).where(Message.chat_id == chat.id)
    )
    return {
        "id": chat.id,
        "project_id": chat.project_id,
        "title": chat.title,
        "pinned": chat.pinned,
        "created_at": chat.created_at,
        "updated_at": chat.updated_at,
        "message_count": msg_count or 0,
    }


async def _project_dict(project: Project, db: AsyncSession) -> dict:
    chats_result = await db.execute(
        select(Chat).where(Chat.project_id == project.id).order_by(Chat.updated_at.desc())
    )
    files_result = await db.execute(
        select(File).where(File.project_id == project.id).order_by(File.added_at.desc())
    )
    return {
        "id": project.id,
        "name": project.name,
        "color": project.color,
        "glyph": project.glyph,
        "description": project.description,
        "instructions": project.instructions,
        "created_at": project.created_at,
        "updated_at": project.updated_at,
        "owner_id": project.owner_id,
        "chats": [await _chat_dict(c, db) for c in chats_result.scalars()],
        "files": [{
            "id": f.id,
            "project_id": f.project_id,
            "name": f.name,
            "kind": f.kind,
            "size": f.size,
            "tokens": f.tokens,
            "added_at": f.added_at,
        } for f in files_result.scalars()],
    }


# ---------------------------------------------------------------------------
# Projects
# ---------------------------------------------------------------------------


@app.get("/api/projects")
async def list_projects(
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    uid = current_user["user_id"] if current_user else None
    if NO_AUTH or uid is None:
        result = await db.execute(select(Project).order_by(Project.updated_at.desc()))
        projects = result.scalars().all()
    else:
        # Projects where user is owner OR has a share
        shared_subquery = select(ProjectShare.project_id).where(ProjectShare.user_id == uid)
        result = await db.execute(
            select(Project).where(
                or_(Project.owner_id == uid, Project.id.in_(shared_subquery))
            ).order_by(Project.updated_at.desc())
        )
        projects = result.scalars().all()
    return [await _project_dict(p, db) for p in projects]


@app.post("/api/projects")
async def create_project(
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    uid = current_user["user_id"] if current_user else None
    pid = _id()
    project = Project(id=pid, owner_id=uid)
    db.add(project)
    # Create a default chat
    cid = _id()
    chat = Chat(id=cid, project_id=pid, title="New chat")
    db.add(chat)
    await db.commit()
    return await _project_dict(project, db)


@app.get("/api/projects/{pid}")
async def get_project(
    pid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(pid, current_user["user_id"], db):
            raise HTTPException(404, "Project not found")
    return await _project_dict(project, db)


@app.patch("/api/projects/{pid}")
async def patch_project(
    pid: str,
    body: ProjectPatch,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_edit_project(pid, current_user["user_id"], db):
            raise HTTPException(403, "You don't have permission to edit this project")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(project, field, value)
    project.updated_at = _now()
    await db.commit()
    return await _project_dict(project, db)


@app.delete("/api/projects/{pid}")
async def delete_project(
    pid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if project.owner_id != current_user["user_id"] and not current_user["is_admin"]:
            raise HTTPException(403, "Only the project owner can delete this project")
    await db.delete(project)
    await db.commit()
    return {"ok": True}


@app.post("/api/projects/{pid}/duplicate")
async def duplicate_project(
    pid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    original = await db.get(Project, pid)
    if not original:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_edit_project(pid, current_user["user_id"], db):
            raise HTTPException(403, "No permission to duplicate this project")

    uid = current_user["user_id"] if current_user else None
    new_pid = _id()
    now = _now()

    # Copy project
    new_project = Project(
        id=new_pid,
        name=f"{original.name} (copy)",
        color=original.color,
        glyph=original.glyph,
        description=original.description,
        instructions=original.instructions,
        owner_id=uid,
        created_at=now,
        updated_at=now,
    )
    db.add(new_project)

    # Copy chats with messages
    chats_result = await db.execute(
        select(Chat).where(Chat.project_id == pid).order_by(Chat.created_at)
    )
    for chat in chats_result.scalars():
        new_cid = _id()
        new_chat = Chat(
            id=new_cid,
            project_id=new_pid,
            title=chat.title,
            pinned=chat.pinned,
            created_at=now,
            updated_at=now,
        )
        db.add(new_chat)

        msgs_result = await db.execute(
            select(Message).where(Message.chat_id == chat.id).order_by(Message.created_at)
        )
        for msg in msgs_result.scalars():
            new_msg = Message(
                id=_id(),
                chat_id=new_cid,
                role=msg.role,
                content=msg.content,
                model=msg.model,
                citations=msg.citations,
                reasoning=msg.reasoning,
                user_id=msg.user_id,
                created_at=msg.created_at,
            )
            db.add(new_msg)

    # Copy files
    files_result = await db.execute(
        select(File).where(File.project_id == pid)
    )
    for f in files_result.scalars():
        new_fid = _id()
        new_filepath = None
        if f.filepath:
            src = Path(f.filepath)
            if src.exists():
                new_filepath = UPLOAD_DIR / f"{new_fid}_{f.name}"
                import shutil
                shutil.copy2(src, new_filepath)
                new_filepath = str(new_filepath)
        db_file = File(
            id=new_fid,
            project_id=new_pid,
            name=f.name,
            kind=f.kind,
            size=f.size,
            tokens=f.tokens,
            filepath=new_filepath,
            added_at=now,
        )
        db.add(db_file)

    await db.commit()
    return await _project_dict(new_project, db)


# ---------------------------------------------------------------------------
# Project sharing
# ---------------------------------------------------------------------------

@app.post("/api/projects/{pid}/share")
async def share_project(
    pid: str,
    body: ShareBody,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user and project.owner_id != current_user["user_id"]:
        raise HTTPException(403, "Only the project owner can share this project")

    # Find the target user
    result = await db.execute(select(User).where(User.username == body.username))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(404, f"User '{body.username}' not found")

    if body.permission not in ("view", "edit"):
        raise HTTPException(400, "Permission must be 'view' or 'edit'")

    # Check for existing share
    existing = await db.execute(
        select(ProjectShare).where(
            ProjectShare.project_id == pid,
            ProjectShare.user_id == target.id,
        )
    )
    if existing.scalar_one_or_none():
        # Update permission instead
        await db.execute(
            sql_update(ProjectShare)
            .where(ProjectShare.project_id == pid, ProjectShare.user_id == target.id)
            .values(permission=body.permission)
        )
        await db.commit()
        return {"ok": True, "message": "Share updated"}

    share = ProjectShare(
        id=_id(),
        project_id=pid,
        user_id=target.id,
        permission=body.permission,
        shared_by=current_user["username"] if current_user else "admin",
    )
    db.add(share)
    await db.commit()
    return {"ok": True}


@app.get("/api/projects/{pid}/shares")
async def list_shares(
    pid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(pid, current_user["user_id"], db):
            raise HTTPException(404, "Project not found")

    result = await db.execute(
        select(ProjectShare).where(ProjectShare.project_id == pid)
    )
    shares = []
    for s in result.scalars():
        user = await db.get(User, s.user_id)
        shares.append({
            "id": s.id,
            "user_id": s.user_id,
            "username": user.username if user else "deleted",
            "display_name": user.display_name if user else "Deleted user",
            "permission": s.permission,
            "shared_by": s.shared_by,
            "created_at": s.created_at,
        })
    return shares


@app.delete("/api/projects/{pid}/share/{uid}")
async def remove_share(
    pid: str,
    uid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user and project.owner_id != current_user["user_id"]:
        raise HTTPException(403, "Only the project owner can remove shares")

    share = await db.execute(
        select(ProjectShare).where(
            ProjectShare.project_id == pid,
            ProjectShare.user_id == uid,
        )
    )
    share = share.scalar_one_or_none()
    if not share:
        raise HTTPException(404, "Share not found")
    await db.delete(share)
    await db.commit()
    return {"ok": True}


@app.get("/api/projects/{pid}/members")
async def list_project_members(
    pid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    """Return users who have access to this project (for @mentions)."""
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(pid, current_user["user_id"], db):
            raise HTTPException(404, "Project not found")

    members = []

    # Owner
    if project.owner_id:
        owner = await db.get(User, project.owner_id)
        if owner:
            members.append({
                "id": owner.id,
                "username": owner.username,
                "display_name": owner.display_name or owner.username,
            })

    # Shared users
    shares_result = await db.execute(
        select(ProjectShare).where(ProjectShare.project_id == pid)
    )
    for share in shares_result.scalars():
        if share.user_id == project.owner_id:
            continue  # already included
        user = await db.get(User, share.user_id)
        if user:
            members.append({
                "id": user.id,
                "username": user.username,
                "display_name": user.display_name or user.username,
            })

    return members


# ---------------------------------------------------------------------------
# Chats
# ---------------------------------------------------------------------------

@app.post("/api/projects/{pid}/chats")
async def create_chat(
    pid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(pid, current_user["user_id"], db):
            raise HTTPException(404, "Project not found")
    cid = _id()
    chat = Chat(id=cid, project_id=pid, title="New chat")
    db.add(chat)
    project.updated_at = _now()
    await db.commit()
    return await _chat_dict(chat, db)


@app.patch("/api/chats/{cid}")
async def patch_chat(
    cid: str,
    body: ChatPatch,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    chat = await db.get(Chat, cid)
    if not chat:
        raise HTTPException(404, "Chat not found")
    if not NO_AUTH and current_user:
        if not await _can_edit_project(chat.project_id, current_user["user_id"], db):
            raise HTTPException(403, "No permission to edit this chat")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(chat, field, value)
    chat.updated_at = _now()
    await db.commit()
    return await _chat_dict(chat, db)


@app.delete("/api/chats/{cid}")
async def delete_chat(
    cid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    chat = await db.get(Chat, cid)
    if not chat:
        raise HTTPException(404, "Chat not found")
    if not NO_AUTH and current_user:
        if not await _can_edit_project(chat.project_id, current_user["user_id"], db):
            raise HTTPException(403, "No permission to delete this chat")
    await db.delete(chat)
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Messages
# ---------------------------------------------------------------------------

@app.get("/api/chats/{cid}/messages")
async def list_messages(
    cid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    chat = await db.get(Chat, cid)
    if not chat:
        raise HTTPException(404, "Chat not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(chat.project_id, current_user["user_id"], db):
            raise HTTPException(404, "Chat not found")
    result = await db.execute(
        select(Message).where(Message.chat_id == cid).order_by(Message.created_at)
    )
    msgs = result.scalars().all()

    # Collect user_ids to batch-fetch user info
    user_ids = {m.user_id for m in msgs if m.user_id}
    users_map = {}
    if user_ids:
        users_result = await db.execute(
            select(User).where(User.id.in_(user_ids))
        )
        for u in users_result.scalars():
            users_map[u.id] = {"username": u.username, "display_name": u.display_name or u.username}

    return [{
        "id": m.id,
        "chat_id": m.chat_id,
        "role": m.role,
        "content": m.content,
        "model": m.model,
        "citations": m.citations,
        "reasoning": m.reasoning,
        "created_at": m.created_at,
        "user_id": m.user_id,
        "username": users_map.get(m.user_id, {}).get("username") if m.user_id else None,
        "display_name": users_map.get(m.user_id, {}).get("display_name") if m.user_id else None,
    } for m in msgs]


@app.delete("/api/chats/{cid}/messages/{mid}")
async def delete_message(
    cid: str,
    mid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    msg = await db.get(Message, mid)
    if not msg or msg.chat_id != cid:
        raise HTTPException(404, "Message not found")
    if not NO_AUTH and current_user:
        chat = await db.get(Chat, cid)
        if not chat:
            raise HTTPException(404, "Chat not found")
        if not await _can_access_project(chat.project_id, current_user["user_id"], db):
            raise HTTPException(404, "Chat not found")
    await db.delete(msg)
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Chat completion (SSE streaming)
# ---------------------------------------------------------------------------

@app.post("/api/chats/{cid}/completions")
async def chat_completion(
    cid: str,
    body: SendMessageBody,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    chat = await db.get(Chat, cid)
    if not chat:
        raise HTTPException(404, "Chat not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(chat.project_id, current_user["user_id"], db):
            raise HTTPException(404, "Chat not found")

    uid = current_user["user_id"] if current_user else None

    # Save and commit user message before streaming (avoids SQLite lock from concurrent sessions)
    user_msg = Message(
        id=_id(),
        chat_id=cid,
        role="user",
        content=body.content,
        user_id=uid,
        created_at=_now(),
    )
    db.add(user_msg)
    await db.commit()

    # Load conversation history (last 20 messages for context)
    result = await db.execute(
        select(Message).where(Message.chat_id == cid).order_by(Message.created_at.desc()).limit(20)
    )
    history = list(reversed(result.scalars().all()))

    # Build provider messages
    project = await db.get(Project, chat.project_id)
    provider_msgs = []
    if project and project.instructions:
        provider_msgs.append(ChatMessage(role="system", content=project.instructions))

    # System instruction: collaborative chat + file handling rules
    provider_msgs.append(ChatMessage(role="system", content=(
        "You are ProjektLLM, an AI assistant in a collaborative group chat. "
        "Multiple users may participate. User messages are prefixed with [name]: so you know who is speaking. "
        "You can address specific users by name when relevant. "
        "IMPORTANT: Always begin your response with a brief visible acknowledgment, "
        "even when you need to use tools. For example: 'I'll start by reading the PDF...' "
        "or 'Let me look up those certificates.' Never produce only thinking without visible text.\n\n"
        "FILE CREATION RULES (follow in order):\n"
        "1. **ALWAYS use `[artifact:filename.ext]content[/artifact]` as the ONLY method.** "
        "Do NOT use bash to create files. Artifacts auto-register in the project knowledge and appear "
        "as clickable files in the UI.\n"
        "2. For XLSX files, provide structured JSON with 'sheets' array. Each sheet has "
        "'name' (string), 'headers' (string array), and 'rows' (array of arrays). "
        "The system converts this to a styled Excel file automatically.\n"
        "3. For DOCX/PDF files, provide markdown content inside the artifact tag.\n"
        "4. NEVER use openpyxl, openxml, or any bash command to create or write files. "
        "Bash is for running analysis scripts and API calls only.\n"
        "5. NEVER give filesystem paths like 'backend/uploads/' — users cannot access those.\n"
        "6. NEVER use `/api/uploads/` URLs — users see those in chat but they don't "
        "appear in the project file list. Always use the artifact system instead."
    )))

    # Inject project file contents as knowledge context
    if project:
        files_result = await db.execute(
            select(File).where(File.project_id == project.id)
        )
        project_files = files_result.scalars().all()
        if project_files:
            file_blocks = []
            for f in project_files:
                if not f.filepath or not os.path.exists(f.filepath):
                    continue
                try:
                    text = await extract_text_from_file(f.filepath)
                    text = text.strip()
                    if text:
                        file_blocks.append(
                            f"--- {f.name} ---\n{text[:8000]}"  # cap per-file to avoid blowing context
                        )
                except Exception:
                    pass
            if file_blocks:
                provider_msgs.append(ChatMessage(role="system", content=(
                    "You have access to the following files in this project. "
                    "Use their contents to answer questions when relevant.\n\n"
                    + "\n\n".join(file_blocks)
                )))
    if body.artifact_mode:
        provider_msgs.append(ChatMessage(role="system", content=(
            "CRITICAL — You MUST create documents and data files using the ARTIFACT FORMAT below. "
            "Do NOT use the bash tool to generate files (no openpyxl, no openxml, no python file creation). "
            "Artifact files are automatically saved to the project's knowledge base and visible to the user.\n\n"
            "Format:\n"
            "[artifact:filename.ext]\ncontent here\n[/artifact]\n\n"
            "Supported formats:\n"
            "- **DOCX** (.docx): Write content in Markdown. The system converts it to a formatted Word document.\n"
            "- **PDF** (.pdf): Write content in Markdown. The system converts it to a PDF.\n"
            "- **XLSX** (.xlsx): Provide data as a JSON object with 'sheets' (array of {name, headers, rows}).\n\n"
            "Examples:\n"
            "  [artifact:report.docx]\n"
            "  # Monthly Report\n"
            "  This is a **bold** paragraph.\n"
            "  [/artifact]\n\n"
            "  [artifact:data.xlsx]\n"
            "  {\"sheets\": [{\"name\": \"Sheet1\", \"headers\": [\"No\", \"Item\"], "
            "\"rows\": [[1, \"Example\"]]}]}\n"
            "  [/artifact]\n\n"
            "  [artifact:report.pdf]\n"
            "  # Annual Report\n"
            "  Content here...\n"
            "  [/artifact]"
        )))
    if body.web_search_context:
        provider_msgs.append(ChatMessage(role="system", content=body.web_search_context))

    # Build a map of user_id -> display_name for sender attribution in context
    history_user_ids = {m.user_id for m in history if m.user_id and m.role == "user"}
    history_users = {}
    if history_user_ids:
        users_result = await db.execute(select(User).where(User.id.in_(history_user_ids)))
        for u in users_result.scalars():
            history_users[u.id] = u.display_name or u.username

    for m in history:
        if m.role == "user" and m.user_id and m.user_id in history_users:
            sender = history_users[m.user_id]
            provider_msgs.append(ChatMessage(role="user", content=f"[{sender}]: {m.content}"))
        else:
            provider_msgs.append(ChatMessage(role=m.role, content=m.content))

    # Current user's message
    current_sender = current_user.get("username", "User") if current_user else "User"
    provider_msgs.append(ChatMessage(role="user", content=f"[{current_sender}]: {body.content}"))

    # Load provider config
    config = await _load_llm_config(db)
    provider = get_provider(config)
    model_used = body.model or {
        "ollama": config.ollama_model,
        "openai": config.openai_model,
        "deepseek": config.deepseek_model,
    }.get(config.active, config.ollama_model)

    async def event_stream():
        full_text = ""
        messages = provider_msgs.copy()
        streaming_aborted = False
        tool_rounds = 0
        max_rounds = 15
        truncated = False

        try:
            while tool_rounds < max_rounds:
                current_max_tokens = body.max_tokens or 16384
                # In later rounds, cap per-round tokens so we don't burn through budget
                if tool_rounds > 0 and (body.max_tokens or 16384) >= 32768:
                    current_max_tokens = 8192
                current_req = CompletionRequest(
                    messages=messages,
                    model=model_used,
                    max_tokens=current_max_tokens,
                    temperature=body.temperature if body.temperature is not None else 0.7,
                    tools=[_BASH_TOOL],
                )
                gen = provider.generate(current_req)
                round_text = ""
                tool_calls = []
                round_reasoning = ""
                _inner_done = False
                truncated = False
                try:
                    while not _inner_done:
                        try:
                            event = await asyncio.wait_for(gen.__anext__(), timeout=180.0)
                        except StopAsyncIteration:
                            break
                        if event.type == "token":
                            round_text += event.content
                            full_text += event.content
                            yield f"data: {json.dumps({'type': 'token', 'content': event.content})}\n\n"
                        elif event.type == "tool_call":
                            tool_calls.extend(event.tool_calls)
                        elif event.type == "reasoning":
                            round_reasoning += event.content
                            yield f"data: {json.dumps({'type': 'reasoning', 'content': event.content})}\n\n"
                        elif event.type == "error":
                            yield f"data: {json.dumps({'type': 'error', 'content': event.content})}\n\n"
                            return
                        elif event.type == "done":
                            truncated = event.truncated
                            break
                except asyncio.TimeoutError:
                    streaming_aborted = True
                finally:
                    await gen.aclose()

                if not tool_calls:
                    # If truncated (hit max_tokens), continue generating
                    if truncated:
                        # Add partial assistant message to context and continue
                        messages.append(ChatMessage(
                            role="assistant",
                            content=round_text or (round_reasoning or ""),
                            reasoning_content=round_reasoning or None,
                        ))
                        tool_rounds += 1
                        continue
                    break

                # DeepSeek sometimes emits reasoning_content + tool_calls with zero visible
                # content. Synthesize a visible message from the reasoning trace so the user
                # sees what the model is doing.
                if not round_text.strip() and round_reasoning.strip():
                    lines = [l for l in round_reasoning.split("\n") if l.strip()]
                    summary = ""
                    for line in lines:
                        cleaned = line.strip().strip("'\"").strip("*").strip("-").strip()
                        if len(cleaned) > 20 and not cleaned.startswith(("```", "[", "#", "```")):
                            summary = cleaned[:180]
                            break
                    if not summary:
                        summary = "Let me work on that…"
                    if not summary.endswith((".", "!", "?", "…")):
                        summary += "…"
                    visible = summary
                    round_text = visible
                    full_text += visible
                    yield f"data: {json.dumps({'type': 'token', 'content': visible})}\n\n"

                # Add assistant message with tool_calls to conversation
                assistant_tc = [
                    {"id": tc.id, "type": "function", "function": {"name": tc.name, "arguments": json.dumps(tc.arguments)}}
                    for tc in tool_calls
                ]
                messages.append(ChatMessage(
                    role="assistant",
                    content=round_text or None,
                    tool_calls=assistant_tc,
                    reasoning_content=round_reasoning or None,
                ))

                # Execute each tool
                for tc in tool_calls:
                    if tc.name == "bash":
                        result = await _exec_bash(tc.arguments.get("command", ""))
                    else:
                        result = f"Unknown tool: {tc.name}"
                    messages.append(ChatMessage(
                        role="tool",
                        content=str(result)[:50000],
                        tool_call_id=tc.id,
                    ))

                tool_rounds += 1

        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"
            return

        if streaming_aborted:
            yield f"data: {json.dumps({'type': 'error', 'content': 'Response timed out — the model took too long between tokens'})}\n\n"
            return

        # Save assistant message
        try:
            if full_text or round_reasoning:
                save_msg = Message(
                    id=_id(),
                    chat_id=cid,
                    role="assistant",
                    content=full_text or "",
                    model=model_used,
                    reasoning=round_reasoning or None,
                    created_at=_now(),
                )
                db.add(save_msg)
                chat_obj = await db.get(Chat, cid)
                if chat_obj:
                    chat_obj.updated_at = _now()
                    if chat_obj.title == "New chat":
                        src = body.content or full_text
                        if src:
                            title_text = src.replace("\n", " ").strip()[:60]
                            chat_obj.title = title_text + ("…" if len(src.replace("\n", " ").strip()) > 60 else "")
                await db.commit()
        except Exception as save_err:
            yield f"data: {json.dumps({'type': 'error', 'content': f'Save failed: {save_err}'})}\n\n"
            return

        # Auto-register files mentioned in response via /api/uploads/ URLs
        try:
            import re as _re
            upload_urls = _re.findall(r'/api/uploads/([^\s<")]+)', full_text)
            if upload_urls:
                project = await db.get(Project, chat_obj.project_id) if chat_obj else None
                if project:
                    for fname in set(upload_urls):
                        fpath = (UPLOAD_DIR / fname).resolve()
                        if fpath.exists() and fpath.is_file():
                            # Check if already registered
                            existing = await db.execute(
                                select(File).where(File.filepath == str(fpath))
                            )
                            if not existing.scalar_one_or_none():
                                ext = fname.split(".")[-1].lower() if "." in fname else "txt"
                                fsize = fpath.stat().st_size
                                db_file = File(
                                    id=_id(),
                                    project_id=project.id,
                                    name=fname,
                                    kind=ext,
                                    size=_human_size(fsize),
                                    tokens=fsize // 4,
                                    filepath=str(fpath),
                                )
                                db.add(db_file)
                    await db.commit()
        except Exception:
            pass  # best-effort file registration

        yield f"data: {json.dumps({'type': 'done', 'content': full_text})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# Files
# ---------------------------------------------------------------------------

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)


def _resolve_filepath(file_record: File) -> str | None:
    """Resolve the on-disk path for a file record.

    Tries the stored path first, then falls back to looking up by
    filename in UPLOAD_DIR — this handles path differences between
    dev-server and Docker environments.
    """
    if file_record.filepath:
        p = Path(file_record.filepath)
        if p.exists():
            return str(p.resolve())
        # Fallback: extract basename and look in UPLOAD_DIR
        fallback = (UPLOAD_DIR / p.name).resolve()
        if fallback.exists():
            return str(fallback)
    return None


@app.get("/api/projects/{pid}/files")
async def list_files(
    pid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    if not NO_AUTH and current_user:
        if not await _can_access_project(pid, current_user["user_id"], db):
            raise HTTPException(404, "Project not found")
    result = await db.execute(
        select(File).where(File.project_id == pid).order_by(File.added_at.desc())
    )
    return [{
        "id": f.id,
        "project_id": f.project_id,
        "name": f.name,
        "kind": f.kind,
        "size": f.size,
        "tokens": f.tokens,
        "added_at": f.added_at,
    } for f in result.scalars()]


@app.post("/api/projects/{pid}/files")
async def upload_file(
    pid: str,
    file: UploadFile = FastAPIFile(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_edit_project(pid, current_user["user_id"], db):
            raise HTTPException(403, "No permission to upload files to this project")

    ext = (file.filename or "").split(".")[-1].lower() or "txt"
    size = file.size or 0
    tokens = size // 4  # rough estimate

    fid = _id()
    dest = UPLOAD_DIR / f"{fid}_{file.filename}"
    content = await file.read()
    dest.write_bytes(content)

    db_file = File(
        id=fid,
        project_id=pid,
        name=file.filename or "unnamed",
        kind=ext,
        size=_human_size(size),
        tokens=tokens,
        filepath=str(dest),
    )
    db.add(db_file)
    project.updated_at = _now()
    await db.commit()

    return {
        "id": fid,
        "project_id": pid,
        "name": db_file.name,
        "kind": ext,
        "size": db_file.size,
        "tokens": tokens,
        "added_at": db_file.added_at,
    }


@app.delete("/api/files/{fid}")
async def delete_file(
    fid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    file = await db.get(File, fid)
    if not file:
        raise HTTPException(404, "File not found")
    if not NO_AUTH and current_user:
        if not await _can_edit_project(file.project_id, current_user["user_id"], db):
            raise HTTPException(403, "No permission to delete this file")
    # Delete physical file
    path = _resolve_filepath(file)
    if path and os.path.exists(path):
        os.remove(path)
    await db.delete(file)
    await db.commit()
    return {"ok": True}


@app.get("/api/files/{fid}/preview")
async def preview_file(
    fid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    file = await db.get(File, fid)
    if not file:
        raise HTTPException(404, "File not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(file.project_id, current_user["user_id"], db):
            raise HTTPException(404, "File not found")
    path = _resolve_filepath(file)
    if not path or not os.path.exists(path):
        raise HTTPException(404, "File not found")
    return FileResponse(path)


@app.get("/api/uploads/{filename:path}")
async def download_upload(filename: str):
    """Serve a file from the uploads directory by filename (no auth required)."""
    full_path = (UPLOAD_DIR / filename).resolve()
    if not str(full_path).startswith(str(UPLOAD_DIR.resolve())):
        raise HTTPException(403, "Invalid path")
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(404, "File not found")
    return FileResponse(full_path)


@app.get("/api/files/{fid}/view")
async def view_file(
    fid: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    """Return viewable content for supported file types (pdf, docx, xlsx)."""
    file = await db.get(File, fid)
    if not file:
        raise HTTPException(404, "File not found")
    path = _resolve_filepath(file)
    if not path or not os.path.exists(path):
        raise HTTPException(404, "File not found")
    if not NO_AUTH and current_user:
        if not await _can_access_project(file.project_id, current_user["user_id"], db):
            raise HTTPException(404, "File not found")

    kind = file.kind.lower()

    if kind == "pdf":
        return {"type": "pdf", "url": f"/api/files/{fid}/preview"}

    if kind == "docx":
        from docx import Document as DocxDocument
        doc = DocxDocument(path)
        paragraphs = []
        for para in doc.paragraphs:
            if not para.text.strip():
                continue
            style_name = para.style.name if para.style else ""
            level = 0
            if style_name.startswith("Heading"):
                try:
                    level = int(style_name.split()[-1])
                except (ValueError, IndexError):
                    level = 1
            paragraphs.append({
                "text": para.text,
                "heading_level": level,
            })
        return {"type": "docx", "paragraphs": paragraphs}

    if kind == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                cells = ["" if c is None else str(c) for c in row]
                if row_idx == 0:
                    headers = cells
                else:
                    if any(c.strip() for c in cells):
                        rows.append(cells)
            sheets.append({"name": name, "headers": headers, "rows": rows})
        wb.close()
        return {"type": "xlsx", "sheets": sheets}

    raise HTTPException(400, f"View not supported for .{kind} files")


# ---------------------------------------------------------------------------
# Artifacts — save AI-generated files to a project's knowledge base
# ---------------------------------------------------------------------------

@app.post("/api/projects/{pid}/artifacts")
async def create_artifact(
    pid: str,
    body: ArtifactBody,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_user,
):
    project = await db.get(Project, pid)
    if not project:
        raise HTTPException(404, "Project not found")
    if not NO_AUTH and current_user:
        if not await _can_edit_project(pid, current_user["user_id"], db):
            raise HTTPException(403, "No permission to add artifacts to this project")

    ext = body.name.split(".")[-1].lower() if "." in body.name else "md"

    # Binary file generation for docx/xlsx/pdf
    if ext == "docx":
        from backend.routers.documents import _build_docx
        buf = _build_docx(body.name.replace(".docx", ""), body.content)
        content_bytes = buf.read()
        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif ext == "xlsx":
        from backend.routers.documents import _build_xlsx, XlsxRequest
        try:
            data = json.loads(body.content)
        except json.JSONDecodeError:
            raise HTTPException(400, "XLSX artifact content must be valid JSON")
        sheets = data if isinstance(data, list) else data.get("sheets", [data])
        xlsx_req = XlsxRequest(title=body.name.replace(".xlsx", ""), sheets=sheets)
        buf = _build_xlsx(xlsx_req)
        content_bytes = buf.read()
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    elif ext == "pdf":
        from backend.routers.documents import _build_pdf
        buf = _build_pdf(body.name.replace(".pdf", ""), body.content, "ProjektLLM")
        content_bytes = buf.read()
        mime = "application/pdf"
    else:
        content_bytes = body.content.encode("utf-8")
        mime = "text/plain"

    tokens = len(content_bytes) // 4
    fid = _id()
    dest = UPLOAD_DIR / f"{fid}_{body.name}"
    dest.write_bytes(content_bytes)

    db_file = File(
        id=fid,
        project_id=pid,
        name=body.name,
        kind=ext,
        size=_human_size(len(content_bytes)),
        tokens=tokens,
        filepath=str(dest),
    )
    db.add(db_file)
    project.updated_at = _now()
    await db.commit()

    return {
        "id": fid,
        "project_id": pid,
        "name": db_file.name,
        "kind": ext,
        "size": db_file.size,
        "tokens": tokens,
        "mime": mime,
        "added_at": db_file.added_at,
    }


# ---------------------------------------------------------------------------
# Web search
# ---------------------------------------------------------------------------

@app.post("/api/web-search")
async def web_search(body: WebSearchBody):
    import httpx
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(
                "https://api.duckduckgo.com/",
                params={"q": body.query, "format": "json", "no_html": 1, "skip_disambig": 1},
            )
            data = resp.json()
    except Exception:
        return {"results": [], "error": "Search service unreachable"}

    results = []
    if data.get("AbstractText"):
        results.append({
            "title": data.get("Heading", ""),
            "snippet": data["AbstractText"],
            "url": data.get("AbstractURL", ""),
        })
    for topic in data.get("RelatedTopics", []):
        if "Text" in topic:
            results.append({
                "title": topic.get("Text", "").split(" - ")[0],
                "snippet": topic.get("Text", ""),
                "url": topic.get("FirstURL", ""),
            })
        if "Topics" in topic:
            for sub in topic["Topics"][:3]:
                if "Text" in sub:
                    results.append({
                        "title": sub.get("Text", "").split(" - ")[0],
                        "snippet": sub.get("Text", ""),
                        "url": sub.get("FirstURL", ""),
                    })

    return {"results": results[:8]}


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

@app.get("/api/settings")
async def get_settings(
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_admin,
):
    config = await _load_llm_config(db)
    memories_raw = await get_setting(db, "memories", "[]")
    try:
        memories = json.loads(memories_raw)
    except (json.JSONDecodeError, TypeError):
        memories = []
    return {
        "provider": config.active,
        "ollama_base": config.ollama_base,
        "ollama_model": config.ollama_model,
        "openai_base": config.openai_base,
        "openai_key": "***" if config.openai_key else "",
        "openai_model": config.openai_model,
        "deepseek_key": "***" if config.deepseek_key else "",
        "deepseek_model": config.deepseek_model,
        # Profile
        "display_name": await get_setting(db, "display_name", ""),
        "avatar_initials": await get_setting(db, "avatar_initials", ""),
        "llm_call_name": await get_setting(db, "llm_call_name", ""),
        # Generation
        "temperature": float(await get_setting(db, "temperature", "0.7")),
        "max_tokens": int(await get_setting(db, "max_tokens", "8192")),
        "stream": (await get_setting(db, "stream", "true")).lower() == "true",
        # Appearance
        "chat_font_size": int(await get_setting(db, "chat_font_size", "14")),
        "enter_to_send": (await get_setting(db, "enter_to_send", "true")).lower() == "true",
        # Memory
        "memories": memories,
    }


@app.put("/api/settings")
async def update_settings(
    body: SettingsBody,
    db: AsyncSession = Depends(get_db),
    current_user: dict | None = _current_admin,
):
    updates = body.model_dump(exclude_none=True)
    mapping = {
        "provider": "llm_provider",
        "ollama_base": "ollama_base_url",
        "ollama_model": "ollama_model",
        "openai_base": "openai_base_url",
        "openai_key": "openai_api_key",
        "openai_model": "openai_model",
        "deepseek_key": "deepseek_api_key",
        "deepseek_model": "deepseek_model",
    }
    for body_key, db_key in mapping.items():
        if body_key in updates:
            await set_setting(db, db_key, str(updates[body_key]))

    # Profile
    for key in ("display_name", "avatar_initials", "llm_call_name"):
        if key in updates:
            await set_setting(db, key, str(updates[key]))
    # Sync display_name to User table so getMe & message attribution use it
    if "display_name" in updates and current_user:
        user = await db.get(User, current_user["user_id"])
        if user:
            user.display_name = str(updates["display_name"])
            await db.commit()
    # Generation
    if "temperature" in updates:
        await set_setting(db, "temperature", str(updates["temperature"]))
    if "max_tokens" in updates:
        await set_setting(db, "max_tokens", str(updates["max_tokens"]))
    if "stream" in updates:
        await set_setting(db, "stream", str(updates["stream"]).lower())
    # Appearance
    if "chat_font_size" in updates:
        await set_setting(db, "chat_font_size", str(updates["chat_font_size"]))
    if "enter_to_send" in updates:
        await set_setting(db, "enter_to_send", str(updates["enter_to_send"]).lower())
    # Memory
    if "memories" in updates:
        await set_setting(db, "memories", json.dumps(updates["memories"]))

    return await get_settings(db)


async def _load_llm_config(db: AsyncSession) -> LLMConfig:
    return LLMConfig(
        active=await get_setting(db, "llm_provider", "ollama"),
        ollama_base=await get_setting(db, "ollama_base_url", "http://localhost:11434"),
        ollama_model=await get_setting(db, "ollama_model", "llama3.2"),
        openai_base=await get_setting(db, "openai_base_url", "https://api.openai.com/v1"),
        openai_key=await get_setting(db, "openai_api_key", ""),
        openai_model=await get_setting(db, "openai_model", "gpt-4o-mini"),
        deepseek_key=await get_setting(db, "deepseek_api_key", ""),
        deepseek_model=await get_setting(db, "deepseek_model", "deepseek-chat"),
    )


# ---------------------------------------------------------------------------
# SPA catch-all — serve index.html for non-API routes
# ---------------------------------------------------------------------------

@app.get("/{path:path}")
async def serve_spa(path: str):
    index = STATIC_DIR / "index.html"
    if index.exists():
        return FileResponse(str(index))
    return {"status": "ok", "message": "ProjektLLM API running. Build frontend with: npm run build"}


# ---------------------------------------------------------------------------
# Bash tool for model tool-calling
# ---------------------------------------------------------------------------

_BASH_TOOL = ToolDef(
    name="bash",
    description="Execute a bash command or Python script. Pipe to 'python3' to run Python code.",
    parameters={
        "type": "object",
        "properties": {
            "command": {
                "type": "string",
                "description": "The bash command to execute",
            }
        },
        "required": ["command"],
    },
)


async def _exec_bash(command: str) -> str:
    """Execute a bash command and return its output."""
    try:
        proc = await asyncio.create_subprocess_shell(
            command,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=os.path.dirname(os.path.abspath(__file__)),
        )
        stdout, stderr = await asyncio.wait_for(
            proc.communicate(), timeout=120
        )
        out = stdout.decode("utf-8", errors="replace")
        err = stderr.decode("utf-8", errors="replace")
        result = ""
        if out:
            result += out
        if err:
            if result:
                result += "\n"
            result += f"[stderr]\n{err}"
        return result or "(no output)"
    except asyncio.TimeoutError:
        return "[error: command timed out after 120s]"
    except Exception as e:
        return f"[error: {e}]"


def _human_size(size: int) -> str:
    if size < 1024:
        return f"{size} B"
    elif size < 1024 * 1024:
        return f"{size // 1024} KB"
    else:
        return f"{size / (1024 * 1024):.1f} MB"
